#!/usr/bin/env python3
"""
Generates DGII-compliant XML test files from the official DGII Excel test set.
Follows the XSD element ordering strictly (xs:sequence).
"""
import openpyxl
import sys
import os
import re
from datetime import datetime
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom

sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r'C:\Users\SAMUE\Downloads\40214827087-21052026010632.xlsx'
OUTPUT_DIR = r'd:\WEBS\GRIDBASE DIGITAL SOLUTIONS\Gridbase Bills\storage\app\dgii_tests'
XSD_DIR = r'D:\WEBS\GRIDBASE DIGITAL SOLUTIONS\Documentación Facturación Electónica'

# Maximum number of items per invoice in the Excel
MAX_ITEMS = 62

def read_excel():
    """Read both ECF and RFCE sheets, returning header->value dicts per row."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    ecf_rows = read_sheet(wb['ECF'])
    rfce_rows = read_sheet(wb['RFCE'])
    
    return ecf_rows, rfce_rows

def read_sheet(ws):
    """Read a worksheet returning list of {header: value} dicts."""
    # Read all headers from row 1
    headers = []
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        headers.append(h if h and h != '#e' else None)
    
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        data = {}
        has_data = False
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col).value
            header = headers[col - 1]
            if val is not None and val != '#e' and header is not None:
                data[header] = str(val).strip()
                has_data = True
        if has_data and 'CasoPrueba' in data:
            rows.append(data)
    
    return rows

def add_element(parent, tag, value):
    """Add a child element only if value is non-empty."""
    if value:
        el = SubElement(parent, tag)
        el.text = str(value)
        return el
    return None

def get(data, key, default=None):
    """Get a value from data dict, return None if empty."""
    v = data.get(key, default)
    if v is None or str(v).strip() == '':
        return None
    return str(v).strip()

def build_ecf_xml(data):
    """Build a complete e-CF XML following XSD element order."""
    tipo = get(data, 'TipoeCF', '31')
    
    root = Element('ECF')
    root.set('xmlns', 'http://dgii.gov.do/e-CF')
    
    enc = SubElement(root, 'Encabezado')
    
    # 1. Version (MUST BE FIRST in Encabezado per XSD)
    add_element(enc, 'Version', get(data, 'Version', '1.0'))
    
    # 2. IdDoc
    iddoc = SubElement(enc, 'IdDoc')
    add_element(iddoc, 'TipoeCF', tipo)
    add_element(iddoc, 'eNCF', get(data, 'ENCF'))
    add_element(iddoc, 'FechaVencimientoSecuencia', get(data, 'FechaVencimientoSecuencia'))
    add_element(iddoc, 'IndicadorNotaCredito', get(data, 'IndicadorNotaCredito'))
    add_element(iddoc, 'IndicadorEnvioDiferido', get(data, 'IndicadorEnvioDiferido'))
    add_element(iddoc, 'IndicadorMontoGravado', get(data, 'IndicadorMontoGravado'))
    add_element(iddoc, 'IndicadorServicioTodoIncluido', get(data, 'IndicadorServicioTodoIncluido'))
    add_element(iddoc, 'TipoIngresos', get(data, 'TipoIngresos'))
    add_element(iddoc, 'TipoPago', get(data, 'TipoPago'))
    add_element(iddoc, 'FechaLimitePago', get(data, 'FechaLimitePago'))
    add_element(iddoc, 'TerminoPago', get(data, 'TerminoPago'))
    
    # TablaFormasPago
    formas = []
    for i in range(1, 8):
        fp = get(data, f'FormaPago[{i}]')
        mp = get(data, f'MontoPago[{i}]')
        if fp and mp:
            formas.append((fp, mp))
    if formas:
        tabla = SubElement(iddoc, 'TablaFormasPago')
        for fp, mp in formas:
            fdp = SubElement(tabla, 'FormaDePago')
            add_element(fdp, 'FormaPago', fp)
            add_element(fdp, 'MontoPago', mp)
    
    add_element(iddoc, 'TipoCuentaPago', get(data, 'TipoCuentaPago'))
    add_element(iddoc, 'NumeroCuentaPago', get(data, 'NumeroCuentaPago'))
    add_element(iddoc, 'BancoPago', get(data, 'BancoPago'))
    add_element(iddoc, 'FechaDesde', get(data, 'FechaDesde'))
    add_element(iddoc, 'FechaHasta', get(data, 'FechaHasta'))
    add_element(iddoc, 'TotalPaginas', get(data, 'TotalPaginas'))
    
    # 3. Emisor
    emisor = SubElement(enc, 'Emisor')
    add_element(emisor, 'RNCEmisor', get(data, 'RNCEmisor'))
    add_element(emisor, 'RazonSocialEmisor', get(data, 'RazonSocialEmisor'))
    add_element(emisor, 'NombreComercial', get(data, 'NombreComercial'))
    add_element(emisor, 'Sucursal', get(data, 'Sucursal'))
    add_element(emisor, 'DireccionEmisor', get(data, 'DireccionEmisor'))
    add_element(emisor, 'Municipio', get(data, 'Municipio'))
    add_element(emisor, 'Provincia', get(data, 'Provincia'))
    
    # TablaTelefonoEmisor
    telefonos = []
    for i in range(1, 4):
        t = get(data, f'TelefonoEmisor[{i}]')
        if t:
            telefonos.append(t)
    if telefonos:
        tabla_tel = SubElement(emisor, 'TablaTelefonoEmisor')
        for t in telefonos:
            add_element(tabla_tel, 'TelefonoEmisor', t)
    
    add_element(emisor, 'CorreoEmisor', get(data, 'CorreoEmisor'))
    add_element(emisor, 'WebSite', get(data, 'WebSite'))
    add_element(emisor, 'ActividadEconomica', get(data, 'ActividadEconomica'))
    add_element(emisor, 'CodigoVendedor', get(data, 'CodigoVendedor'))
    add_element(emisor, 'NumeroFacturaInterna', get(data, 'NumeroFacturaInterna'))
    add_element(emisor, 'NumeroPedidoInterno', get(data, 'NumeroPedidoInterno'))
    add_element(emisor, 'ZonaVenta', get(data, 'ZonaVenta'))
    add_element(emisor, 'RutaVenta', get(data, 'RutaVenta'))
    add_element(emisor, 'InformacionAdicionalEmisor', get(data, 'InformacionAdicionalEmisor'))
    add_element(emisor, 'FechaEmision', get(data, 'FechaEmision'))
    
    # 4. Comprador
    comprador_fields = [
        'RNCComprador', 'IdentificadorExtranjero', 'RazonSocialComprador',
        'ContactoComprador', 'CorreoComprador', 'DireccionComprador',
        'MunicipioComprador', 'ProvinciaComprador', 'PaisComprador',
        'FechaEntrega', 'ContactoEntrega', 'DireccionEntrega',
        'TelefonoAdicional', 'FechaOrdenCompra', 'NumeroOrdenCompra',
        'CodigoInternoComprador', 'ResponsablePago', 'InformacionAdicionalComprador'
    ]
    has_comprador = any(get(data, f) for f in comprador_fields)
    if has_comprador:
        comprador = SubElement(enc, 'Comprador')
        for f in comprador_fields:
            add_element(comprador, f, get(data, f))
    
    # 5. InformacionesAdicionales
    info_add_fields = [
        'FechaEmbarque', 'NumeroEmbarque', 'NumeroContenedor',
        'NumeroReferencia', 'NombrePuertoEmbarque', 'CondicionesEntrega',
        'TotalFob', 'Seguro', 'Flete', 'OtrosGastos', 'TotalCif',
        'RegimenAduanero', 'NombrePuertoSalida', 'NombrePuertoDesembarque',
        'PesoBruto', 'PesoNeto', 'UnidadPesoBruto', 'UnidadPesoNeto',
        'CantidadBulto', 'UnidadBulto', 'VolumenBulto', 'UnidadVolumen'
    ]
    # Check for field with trailing space too
    has_info = any(get(data, f) or get(data, f + ' ') for f in info_add_fields)
    if has_info:
        info_add = SubElement(enc, 'InformacionesAdicionales')
        for f in info_add_fields:
            val = get(data, f) or get(data, f + ' ')  # Handle trailing space in header
            add_element(info_add, f.strip(), val)
    
    # 6. Transporte
    transport_fields = [
        'ViaTransporte', 'PaisOrigen', 'DireccionDestino', 'PaisDestino',
        'RNCIdentificacionCompaniaTransportista', 'NombreCompaniaTransportista',
        'NumeroViaje', 'Conductor', 'DocumentoTransporte', 'Ficha', 'Placa',
        'RutaTransporte', 'ZonaTransporte', 'NumeroAlbaran'
    ]
    has_transport = any(get(data, f) for f in transport_fields)
    if has_transport:
        transporte = SubElement(enc, 'Transporte')
        for f in transport_fields:
            add_element(transporte, f, get(data, f))
    
    # 7. Totales
    totales = SubElement(enc, 'Totales')
    add_element(totales, 'MontoGravadoTotal', get(data, 'MontoGravadoTotal'))
    add_element(totales, 'MontoGravadoI1', get(data, 'MontoGravadoI1'))
    add_element(totales, 'MontoGravadoI2', get(data, 'MontoGravadoI2'))
    add_element(totales, 'MontoGravadoI3', get(data, 'MontoGravadoI3'))
    add_element(totales, 'MontoExento', get(data, 'MontoExento'))
    add_element(totales, 'ITBIS1', get(data, 'ITBIS1'))
    add_element(totales, 'ITBIS2', get(data, 'ITBIS2'))
    add_element(totales, 'ITBIS3', get(data, 'ITBIS3'))
    add_element(totales, 'TotalITBIS', get(data, 'TotalITBIS'))
    add_element(totales, 'TotalITBIS1', get(data, 'TotalITBIS1'))
    add_element(totales, 'TotalITBIS2', get(data, 'TotalITBIS2'))
    add_element(totales, 'TotalITBIS3', get(data, 'TotalITBIS3'))
    add_element(totales, 'MontoImpuestoAdicional', get(data, 'MontoImpuestoAdicional'))
    
    # ImpuestosAdicionales in Totales
    imp_tipos = []
    for i in range(1, 5):
        tip = get(data, f'TipoImpuesto[{i}]')
        if tip:
            imp = {'TipoImpuesto': tip}
            for f in ['TasaImpuestoAdicional', 'MontoImpuestoSelectivoConsumoEspecifico',
                       'MontoImpuestoSelectivoConsumoAdvalorem', 'OtrosImpuestosAdicionales']:
                v = get(data, f'{f}[{i}]')
                if v:
                    imp[f] = v
            imp_tipos.append(imp)
    if imp_tipos:
        imps = SubElement(totales, 'ImpuestosAdicionales')
        for imp in imp_tipos:
            ia = SubElement(imps, 'ImpuestoAdicional')
            add_element(ia, 'TipoImpuesto', imp['TipoImpuesto'])
            for f in ['TasaImpuestoAdicional', 'MontoImpuestoSelectivoConsumoEspecifico',
                       'MontoImpuestoSelectivoConsumoAdvalorem', 'OtrosImpuestosAdicionales']:
                add_element(ia, f, imp.get(f))
    
    add_element(totales, 'MontoTotal', get(data, 'MontoTotal'))
    add_element(totales, 'MontoNoFacturable', get(data, 'MontoNoFacturable'))
    add_element(totales, 'MontoPeriodo', get(data, 'MontoPeriodo'))
    add_element(totales, 'SaldoAnterior', get(data, 'SaldoAnterior'))
    add_element(totales, 'MontoAvancePago', get(data, 'MontoAvancePago'))
    add_element(totales, 'ValorPagar', get(data, 'ValorPagar'))
    add_element(totales, 'TotalITBISRetenido', get(data, 'TotalITBISRetenido'))
    add_element(totales, 'TotalISRRetencion', get(data, 'TotalISRRetencion'))
    add_element(totales, 'TotalITBISPercepcion', get(data, 'TotalITBISPercepcion'))
    add_element(totales, 'TotalISRPercepcion', get(data, 'TotalISRPercepcion'))
    
    # 8. OtraMoneda
    otra_fields = ['TipoMoneda', 'TipoCambio', 'MontoGravadoTotalOtraMoneda',
                   'MontoGravado1OtraMoneda', 'MontoGravado2OtraMoneda', 'MontoGravado3OtraMoneda',
                   'MontoExentoOtraMoneda', 'TotalITBISOtraMoneda', 'TotalITBIS1OtraMoneda',
                   'TotalITBIS2OtraMoneda', 'TotalITBIS3OtraMoneda',
                   'MontoImpuestoAdicionalOtraMoneda', 'MontoTotalOtraMoneda']
    has_otra = any(get(data, f) for f in otra_fields)
    if has_otra:
        otra = SubElement(enc, 'OtraMoneda')
        for f in otra_fields:
            add_element(otra, f, get(data, f))
    
    # === DetallesItems ===
    items_section = SubElement(root, 'DetallesItems')
    for n in range(1, MAX_ITEMS + 1):
        nl = get(data, f'NumeroLinea[{n}]')
        if not nl:
            continue
        
        item = SubElement(items_section, 'Item')
        add_element(item, 'NumeroLinea', nl)
        
        # TablaCodigosItem
        codigos = []
        for ci in range(1, 6):
            tc = get(data, f'TipoCodigo[{n}][{ci}]')
            cc = get(data, f'CodigoItem[{n}][{ci}]')
            if tc and cc:
                codigos.append((tc, cc))
        if codigos:
            tabla_cod = SubElement(item, 'TablaCodigosItem')
            for tc, cc in codigos:
                ci_el = SubElement(tabla_cod, 'CodigosItem')
                add_element(ci_el, 'TipoCodigo', tc)
                add_element(ci_el, 'CodigoItem', cc)
        
        add_element(item, 'IndicadorFacturacion', get(data, f'IndicadorFacturacion[{n}]'))
        
        # Retencion
        ret_fields = ['IndicadorAgenteRetencionoPercepcion', 'MontoITBISRetenido', 'MontoISRRetenido']
        has_ret = any(get(data, f'{f}[{n}]') for f in ret_fields)
        if has_ret:
            ret = SubElement(item, 'Retencion')
            for f in ret_fields:
                add_element(ret, f, get(data, f'{f}[{n}]'))
        
        add_element(item, 'NombreItem', get(data, f'NombreItem[{n}]'))
        add_element(item, 'IndicadorBienoServicio', get(data, f'IndicadorBienoServicio[{n}]'))
        add_element(item, 'DescripcionItem', get(data, f'DescripcionItem[{n}]'))
        add_element(item, 'CantidadItem', get(data, f'CantidadItem[{n}]'))
        add_element(item, 'UnidadMedida', get(data, f'UnidadMedida[{n}]'))
        add_element(item, 'CantidadReferencia', get(data, f'CantidadReferencia[{n}]'))
        add_element(item, 'UnidadReferencia', get(data, f'UnidadReferencia[{n}]'))
        
        # TablaSubcantidad
        subcants = []
        for si in range(1, 6):
            sc = get(data, f'Subcantidad[{n}][{si}]')
            cs = get(data, f'CodigoSubcantidad[{n}][{si}]')
            if sc or cs:
                subcants.append((sc, cs))
        if subcants:
            tabla_sub = SubElement(item, 'TablaSubcantidad')
            for sc, cs in subcants:
                sci = SubElement(tabla_sub, 'SubcantidadItem')
                add_element(sci, 'Subcantidad', sc)
                add_element(sci, 'CodigoSubcantidad', cs)
        
        add_element(item, 'GradosAlcohol', get(data, f'GradosAlcohol[{n}]'))
        add_element(item, 'PrecioUnitarioReferencia', get(data, f'PrecioUnitarioReferencia[{n}]'))
        add_element(item, 'FechaElaboracion', get(data, f'FechaElaboracion[{n}]'))
        add_element(item, 'FechaVencimientoItem', get(data, f'FechaVencimientoItem[{n}]'))
        
        # Mineria
        min_fields = ['PesoNetoKilogramo', 'PesoNetoMineria', 'TipoAfiliacion', 'Liquidacion']
        has_min = any(get(data, f'{f}[{n}]') for f in min_fields)
        if has_min:
            mineria = SubElement(item, 'Mineria')
            for f in min_fields:
                add_element(mineria, f, get(data, f'{f}[{n}]'))
        
        add_element(item, 'PrecioUnitarioItem', get(data, f'PrecioUnitarioItem[{n}]'))
        add_element(item, 'DescuentoMonto', get(data, f'DescuentoMonto[{n}]'))
        
        # TablaSubDescuento
        subdesc = []
        for si in range(1, 6):
            tsd = get(data, f'TipoSubDescuento[{n}][{si}]')
            if tsd:
                subdesc.append({
                    'TipoSubDescuento': tsd,
                    'SubDescuentoPorcentaje': get(data, f'SubDescuentoPorcentaje[{n}][{si}]'),
                    'MontoSubDescuento': get(data, f'MontoSubDescuento[{n}][{si}]'),
                })
        if subdesc:
            tabla_sd = SubElement(item, 'TablaSubDescuento')
            for sd in subdesc:
                sde = SubElement(tabla_sd, 'SubDescuento')
                add_element(sde, 'TipoSubDescuento', sd['TipoSubDescuento'])
                add_element(sde, 'SubDescuentoPorcentaje', sd.get('SubDescuentoPorcentaje'))
                add_element(sde, 'MontoSubDescuento', sd.get('MontoSubDescuento'))
        
        add_element(item, 'RecargoMonto', get(data, f'RecargoMonto[{n}]'))
        
        # TablaSubRecargo
        subrec = []
        for si in range(1, 6):
            tsr = get(data, f'TipoSubRecargo[{n}][{si}]')
            if tsr:
                subrec.append({
                    'TipoSubRecargo': tsr,
                    'SubRecargoPorcentaje': get(data, f'SubRecargoPorcentaje[{n}][{si}]'),
                    'MontosubRecargo': get(data, f'MontosubRecargo[{n}][{si}]'),
                })
        if subrec:
            tabla_sr = SubElement(item, 'TablaSubRecargo')
            for sr in subrec:
                sre = SubElement(tabla_sr, 'SubRecargo')
                add_element(sre, 'TipoSubRecargo', sr['TipoSubRecargo'])
                add_element(sre, 'SubRecargoPorcentaje', sr.get('SubRecargoPorcentaje'))
                add_element(sre, 'MontoSubRecargo', sr.get('MontosubRecargo'))
        
        # TablaImpuestoAdicional
        item_imps = []
        for ii in range(1, 3):
            ti = get(data, f'TipoImpuesto[{n}][{ii}]')
            if ti:
                item_imps.append(ti)
        if item_imps:
            tabla_ia = SubElement(item, 'TablaImpuestoAdicional')
            for ti in item_imps:
                ia = SubElement(tabla_ia, 'ImpuestoAdicional')
                add_element(ia, 'TipoImpuesto', ti)
        
        # OtraMonedaDetalle
        omd_fields = ['PrecioOtraMoneda', 'DescuentoOtraMoneda', 'RecargoOtraMoneda', 'MontoItemOtraMoneda']
        has_omd = any(get(data, f'{f}[{n}]') for f in omd_fields)
        if has_omd:
            omd = SubElement(item, 'OtraMonedaDetalle')
            for f in omd_fields:
                add_element(omd, f, get(data, f'{f}[{n}]'))
        
        add_element(item, 'MontoItem', get(data, f'MontoItem[{n}]'))
    
    # === Subtotales ===
    st_num = get(data, 'NumeroSubTotal')
    if st_num:
        subtotales = SubElement(root, 'Subtotales')
        st = SubElement(subtotales, 'Subtotal')
        add_element(st, 'NumeroSubTotal', st_num)
        add_element(st, 'DescripcionSubtotal', get(data, 'DescripcionSubtotal'))
        add_element(st, 'Orden', get(data, 'Orden'))
        for f in ['SubTotalMontoGravadoTotal', 'SubTotalMontoGravadoI1', 'SubTotalMontoGravadoI2',
                   'SubTotalMontoGravadoI3', 'SubTotaITBIS', 'SubTotaITBIS1', 'SubTotaITBIS2',
                   'SubTotaITBIS3', 'SubTotalImpuestoAdicional', 'SubTotalExento',
                   'MontoSubTotal', 'Lineas']:
            add_element(st, f, get(data, f))
    
    # === DescuentosORecargos ===
    dor_items = []
    for i in range(1, 3):
        nl_dor = get(data, f'NumeroLineaDoR[{i}]')
        if nl_dor:
            dor = {'NumeroLinea': nl_dor}
            for f in ['TipoAjuste', 'IndicadorNorma1007', 'DescripcionDescuentooRecargo',
                       'TipoValor', 'ValorDescuentooRecargo', 'MontoDescuentooRecargo',
                       'MontoDescuentooRecargoOtraMoneda', 'IndicadorFacturacionDescuentooRecargo']:
                dor[f] = get(data, f'{f}[{i}]')
            dor_items.append(dor)
    if dor_items:
        dors = SubElement(root, 'DescuentosORecargos')
        for dor in dor_items:
            dore = SubElement(dors, 'DescuentoORecargo')
            add_element(dore, 'NumeroLinea', dor['NumeroLinea'])
            for f in ['TipoAjuste', 'IndicadorNorma1007', 'DescripcionDescuentooRecargo',
                       'TipoValor', 'ValorDescuentooRecargo', 'MontoDescuentooRecargo',
                       'MontoDescuentooRecargoOtraMoneda', 'IndicadorFacturacionDescuentooRecargo']:
                add_element(dore, f, dor.get(f))
    
    # === Paginacion ===
    pag_items = []
    for i in range(1, 3):
        pn = get(data, f'PaginaNo[{i}]')
        if pn:
            pag = {'PaginaNo': pn}
            for f in ['NoLineaDesde', 'NoLineaHasta', 'SubtotalMontoGravadoPagina',
                       'SubtotalMontoGravado1Pagina', 'SubtotalMontoGravado2Pagina',
                       'SubtotalMontoGravado3Pagina', 'SubtotalExentoPagina',
                       'SubtotalItbisPagina', 'SubtotalItbis1Pagina', 'SubtotalItbis2Pagina',
                       'SubtotalItbis3Pagina', 'SubtotalImpuestoAdicionalPagina',
                       'MontoSubtotalPagina', 'SubtotalMontoNoFacturablePagina']:
                pag[f] = get(data, f'{f}[{i}]')
            pag_items.append(pag)
    if pag_items:
        pags = SubElement(root, 'Paginacion')
        for pag in pag_items:
            pe = SubElement(pags, 'Pagina')
            add_element(pe, 'PaginaNo', pag['PaginaNo'])
            for f in ['NoLineaDesde', 'NoLineaHasta', 'SubtotalMontoGravadoPagina',
                       'SubtotalMontoGravado1Pagina', 'SubtotalMontoGravado2Pagina',
                       'SubtotalMontoGravado3Pagina', 'SubtotalExentoPagina',
                       'SubtotalItbisPagina', 'SubtotalItbis1Pagina', 'SubtotalItbis2Pagina',
                       'SubtotalItbis3Pagina', 'SubtotalImpuestoAdicionalPagina',
                       'MontoSubtotalPagina', 'SubtotalMontoNoFacturablePagina']:
                add_element(pe, f, pag.get(f))
    
    # === InformacionReferencia === (required for types 33, 34)
    ncf_mod = get(data, 'NCFModificado')
    if ncf_mod:
        info_ref = SubElement(root, 'InformacionReferencia')
        add_element(info_ref, 'NCFModificado', ncf_mod)
        add_element(info_ref, 'RNCOtroContribuyente', get(data, 'RNCOtroContribuyente'))
        add_element(info_ref, 'FechaNCFModificado', get(data, 'FechaNCFModificado'))
        add_element(info_ref, 'CodigoModificacion', get(data, 'CodigoModificacion'))
        add_element(info_ref, 'RazonModificacion', get(data, 'RazonModificacion'))
    
    # === FechaHoraFirma === (DGII pattern: dd-MM-yyyy HH:mm:ss)
    add_element(root, 'FechaHoraFirma', datetime.now().strftime('%d-%m-%Y %H:%M:%S'))
    
    return root


def build_rfce_xml(data):
    """Build an RFCE (Resumen Factura Consumo Electronica) XML."""
    root = Element('RFCE')
    root.set('xmlns', 'http://dgii.gov.do/RFCE')
    
    enc = SubElement(root, 'Encabezado')
    
    # Version first
    add_element(enc, 'Version', get(data, 'Version', '1.0'))
    
    # IdDoc
    iddoc = SubElement(enc, 'IdDoc')
    add_element(iddoc, 'TipoeCF', get(data, 'TipoeCF', '32'))
    add_element(iddoc, 'eNCF', get(data, 'ENCF'))
    add_element(iddoc, 'TipoIngresos', get(data, 'TipoIngresos'))
    add_element(iddoc, 'TipoPago', get(data, 'TipoPago'))
    
    # TablaFormasPago
    formas = []
    for i in range(1, 8):
        fp = get(data, f'FormaPago[{i}]')
        mp = get(data, f'MontoPago[{i}]')
        if fp and mp:
            formas.append((fp, mp))
    if formas:
        tabla = SubElement(iddoc, 'TablaFormasPago')
        for fp, mp in formas:
            fdp = SubElement(tabla, 'FormaDePago')
            add_element(fdp, 'FormaPago', fp)
            add_element(fdp, 'MontoPago', mp)
    
    # Emisor
    emisor = SubElement(enc, 'Emisor')
    add_element(emisor, 'RNCEmisor', get(data, 'RNCEmisor'))
    add_element(emisor, 'RazonSocialEmisor', get(data, 'RazonSocialEmisor'))
    add_element(emisor, 'FechaEmision', get(data, 'FechaEmision'))
    
    # Comprador (optional for RFCE)
    rnc_comp = get(data, 'RNCComprador')
    id_ext = get(data, 'IdentificadorExtranjero')
    razon_comp = get(data, 'RazonSocialComprador')
    if rnc_comp or id_ext or razon_comp:
        comprador = SubElement(enc, 'Comprador')
        add_element(comprador, 'RNCComprador', rnc_comp)
        add_element(comprador, 'IdentificadorExtranjero', id_ext)
        add_element(comprador, 'RazonSocialComprador', razon_comp)
    
    # Totales
    totales = SubElement(enc, 'Totales')
    add_element(totales, 'MontoGravadoTotal', get(data, 'MontoGravadoTotal'))
    add_element(totales, 'MontoGravadoI1', get(data, 'MontoGravadoI1'))
    add_element(totales, 'MontoGravadoI2', get(data, 'MontoGravadoI2'))
    add_element(totales, 'MontoGravadoI3', get(data, 'MontoGravadoI3'))
    add_element(totales, 'MontoExento', get(data, 'MontoExento'))
    add_element(totales, 'TotalITBIS', get(data, 'TotalITBIS'))
    add_element(totales, 'TotalITBIS1', get(data, 'TotalITBIS1'))
    add_element(totales, 'TotalITBIS2', get(data, 'TotalITBIS2'))
    add_element(totales, 'TotalITBIS3', get(data, 'TotalITBIS3'))
    add_element(totales, 'MontoImpuestoAdicional', get(data, 'MontoImpuestoAdicional'))
    
    # ImpuestosAdicionales
    imp_tipos = []
    for i in range(1, 5):
        tip = get(data, f'TipoImpuesto[{i}]')
        if tip:
            imp = {'TipoImpuesto': tip}
            for f in ['MontoImpuestoSelectivoConsumoEspecifico',
                       'MontoImpuestoSelectivoConsumoAdvalorem', 'OtrosImpuestosAdicionales']:
                v = get(data, f'{f}[{i}]')
                if v:
                    imp[f] = v
            imp_tipos.append(imp)
    if imp_tipos:
        imps = SubElement(totales, 'ImpuestosAdicionales')
        for imp in imp_tipos:
            ia = SubElement(imps, 'ImpuestoAdicional')
            add_element(ia, 'TipoImpuesto', imp['TipoImpuesto'])
            for f in ['MontoImpuestoSelectivoConsumoEspecifico',
                       'MontoImpuestoSelectivoConsumoAdvalorem', 'OtrosImpuestosAdicionales']:
                add_element(ia, f, imp.get(f))
    
    add_element(totales, 'MontoTotal', get(data, 'MontoTotal'))
    add_element(totales, 'MontoNoFacturable', get(data, 'MontoNoFacturable'))
    add_element(totales, 'MontoPeriodo', get(data, 'MontoPeriodo'))
    
    # CodigoSeguridadeCF
    add_element(root, 'CodigoSeguridadeCF', get(data, 'CodigoSeguridadeCF'))
    
    # FechaHoraFirma (DGII pattern: dd-MM-yyyy HH:mm:ss)
    add_element(root, 'FechaHoraFirma', datetime.now().strftime('%d-%m-%Y %H:%M:%S'))
    
    return root


def xml_to_string(root):
    """Convert ElementTree to clean XML string."""
    rough = tostring(root, encoding='unicode', xml_declaration=False)
    return '<?xml version="1.0" encoding="utf-8"?>\n' + rough


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    # Clean old XMLs
    for f in os.listdir(OUTPUT_DIR):
        if f.endswith('.xml'):
            os.remove(os.path.join(OUTPUT_DIR, f))
    
    ecf_rows, rfce_rows = read_excel()
    
    print(f"ECF rows: {len(ecf_rows)}, RFCE rows: {len(rfce_rows)}")
    
    # Generate ECF XMLs
    for i, data in enumerate(ecf_rows, 1):
        encf = get(data, 'ENCF', f'UNKNOWN_{i}')
        tipo = get(data, 'TipoeCF', '??')
        filename = f"ecf_{i:02d}.xml"
        
        root = build_ecf_xml(data)
        xml_str = xml_to_string(root)
        
        filepath = os.path.join(OUTPUT_DIR, filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(xml_str)
        
        print(f"  Generated {filename} (Type {tipo}, {encf})")
    
    # Generate RFCE XMLs
    for i, data in enumerate(rfce_rows, 1):
        encf = get(data, 'ENCF', f'UNKNOWN_{i}')
        filename = f"rfce_{i:02d}.xml"
        
        root = build_rfce_xml(data)
        xml_str = xml_to_string(root)
        
        filepath = os.path.join(OUTPUT_DIR, filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(xml_str)
        
        print(f"  Generated {filename} ({encf})")
    
    print(f"\nDone! Generated {len(ecf_rows) + len(rfce_rows)} XMLs in {OUTPUT_DIR}")


if __name__ == '__main__':
    main()
